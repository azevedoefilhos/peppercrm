from cache_helpers import cache_clientes, cache_fornecedores, cache_categorias, cache_produtos_fornecedor
# concorrentes.py -- PepperCRM
# 2 abas: Marcas concorrentes | Produtos e relacoes (unificada)

import streamlit as st
import pandas as pd
from database import conectar, query


def _ir(p):
    st.session_state["pagina"] = p
    st.session_state["_scroll_topo"] = True
    st.rerun()


def tela_concorrentes():
    st.header("Concorrentes")
    st.caption("Cadastre aqui antes de ir a campo. Alimenta automaticamente a pesquisa de preços.")
    if st.button("⬅ Voltar"):
        _ir("home")

    ABAS_CC = {
        "marcas":  "Marcas concorrentes",
        "prods":   "Produtos e relações",
        "ean":     "🔍 Busca por EAN",
        "import":  "📥 Importar Excel",
    }
    if "cc_aba" not in st.session_state: st.session_state["cc_aba"] = "marcas"
    cols = st.columns(4)
    for col,(k,v) in zip(cols, ABAS_CC.items()):
        ativa = st.session_state["cc_aba"] == k
        if col.button(v, key=f"ccnav_{k}", width="stretch",
                      type="primary" if ativa else "secondary"):
            st.session_state["cc_aba"] = k; st.rerun()
    st.divider()
    a = st.session_state["cc_aba"]
    if a=="marcas":  _marcas()
    elif a=="prods": _produtos_e_relacoes()
    elif a=="ean":   _gestao_por_ean()
    elif a=="import":_importar_concorrentes()


# ==============================================================
# ABA 1 -- MARCAS CONCORRENTES (inalterada)
# ==============================================================

def _marcas():
    import io
    forns = cache_fornecedores()
    if not forns:
        st.warning("Cadastre ao menos um fornecedor antes de cadastrar concorrentes.")
        return

    if st.button("➕ Nova marca concorrente", type="primary", key="btn_nova_marca"):
        st.session_state["show_form_marca"] = True
        st.session_state.pop("marca_editar_id", None)
        st.session_state.pop("marca_confirmar_excluir", None)

    if st.session_state.get("show_form_marca"):
        _form_nova_marca(forns)

    st.divider()

    # ── Filtro por fornecedor ─────────────────────────
    forn_opts = [(None, "— Todos os fornecedores")] + [(f[0], f[1]) for f in forns]
    fil_forn  = st.selectbox("Filtrar por fornecedor", forn_opts,
                             format_func=lambda x: x[1], key="marc_fil_forn")

    dados = query("""
        SELECT conc.concorrente_id, conc.marca_concorrente,
               f.nome_fantasia, conc.origem_cidade, conc.ativo,
               COALESCE(conc.importada,0), conc.importado_por
        FROM concorrente conc
        JOIN fornecedor f ON conc.fornecedor_id=f.fornecedor_id
        ORDER BY f.nome_fantasia, conc.marca_concorrente
    """)

    if not dados:
        st.info("Nenhuma marca cadastrada ainda.")
        return

    # Aplica filtro
    if fil_forn[0]:
        dados_fil = [r for r in dados if r[2] == fil_forn[1]]
    else:
        dados_fil = list(dados)

    # ── Exportar ──────────────────────────────────────
    col_info, col_exp = st.columns([3, 1])
    col_info.caption(f"{len(dados_fil)} marca(s) exibida(s)"
                     + (f" de {len(dados)} total" if fil_forn[0] else ""))
    with col_exp:
        if dados_fil:
            df_exp = pd.DataFrame(
                [(r[1], r[2], r[3] or "—", "Ativo" if r[4] else "Inativo")
                 for r in dados_fil],
                columns=["Marca","Fornecedor ref.","Origem","Status"]
            )
            buf = io.BytesIO()
            df_exp.to_excel(buf, index=False, sheet_name="Marcas concorrentes")
            buf.seek(0)
            st.download_button("⬇️ Exportar Excel", data=buf,
                               file_name="marcas_concorrentes.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               width="stretch")

    if not dados_fil:
        st.info("Nenhuma marca encontrada para o filtro selecionado.")
        return

    h1, h2, h3, h4, h5 = st.columns([2, 2, 1.5, 0.5, 2])
    for col, txt in zip([h1,h2,h3,h4,h5], ["Marca","Fornecedor ref.","Origem","St.",""]):
        col.markdown(f"<small><b>{txt}</b></small>", unsafe_allow_html=True)

    for row in dados_fil:
        cid, marca, forn_n, origem, ativo, _imp, _iby = row
        icone = "✅" if ativo else "❌"
        c1, c2, c3, c4, c5 = st.columns([2, 2, 1.5, 0.5, 2])
        c1.write(marca)
        c2.caption(forn_n)
        _orig_label = origem or "—"
        if _imp:
            _orig_label = f"🌍 {_orig_label}"
            if _iby: _orig_label += f" · via {_iby}"
        c3.caption(_orig_label)
        c4.caption(icone)
        with c5:
            b1, b2, b3 = st.columns(3)
            with b1:
                if st.button("✏️", key=f"ed_{cid}", help="Editar", width="stretch"):
                    st.session_state["marca_editar_id"] = cid
                    st.session_state.pop("marca_confirmar_excluir", None)
                    st.session_state.pop("show_form_marca", None)
                    st.rerun()
            with b2:
                if st.button("🗑️", key=f"ex_{cid}", help="Excluir", width="stretch"):
                    st.session_state["marca_confirmar_excluir"] = cid
                    st.session_state.pop("marca_editar_id", None)
                    st.session_state.pop("show_form_marca", None)
                    st.rerun()
            with b3:
                tip = "Desativar" if ativo else "Reativar"
                lbl = "🔇" if ativo else "🔔"
                if st.button(lbl, key=f"tog_{cid}", help=tip, width="stretch"):
                    conn = conectar()
                    conn.execute("UPDATE concorrente SET ativo=? WHERE concorrente_id=?",
                                 (0 if ativo else 1, cid))
                    conn.commit(); conn.close(); st.rerun()

        if st.session_state.get("marca_editar_id") == cid:
            _form_editar_marca(cid, marca, origem, forns)
        if st.session_state.get("marca_confirmar_excluir") == cid:
            _confirmacao_excluir_marca(cid, marca)


def _form_nova_marca(forns):
    with st.container():
        with st.form("nova_marca", clear_on_submit=True):
            col1, col2 = st.columns(2)
            with col1:
                forn_sel   = st.selectbox("Fornecedor de referência *", forns,
                                          format_func=lambda x: x[1], key="nm_forn",
                                          help="Com qual dos seus fornecedores esta marca compete?")
                marca_n    = st.text_input("Nome da marca *",
                                           placeholder="Ex: Castelo, Heinz, Ponti...")
                origem_n   = st.text_input("País / cidade de origem",
                                           placeholder="Ex: Itália, São Paulo...")
            with col2:
                importada  = st.checkbox("🌍 Marca importada")
                import_por = st.text_input("Importado por",
                                           placeholder="Ex: La Pastina, BRF... (deixe vazio se não aplicável)",
                                           help="Empresa nacional que importa e distribui esta marca")
                obs_n      = st.text_input("Observação")

            col_s, col_c = st.columns(2)
            with col_s: salvar   = st.form_submit_button("💾 Salvar marca", type="primary")
            with col_c: cancelar = st.form_submit_button("Cancelar")

        if salvar:
            _mn = marca_n.strip() if marca_n else ""
            if not _mn:
                st.error("Nome da marca é obrigatório.")
            else:
                dup = query("""SELECT concorrente_id FROM concorrente
                    WHERE LOWER(marca_concorrente)=LOWER(?) AND fornecedor_id=? AND ativo=1""",
                    (_mn, forn_sel[0]))
                if dup:
                    st.error(
                        f"⚠️ A marca **{_mn}** já está cadastrada para este fornecedor. "
                        "Verifique a lista acima ou edite a existente.")
                else:
                    conn = conectar()
                    conn.execute("""INSERT INTO concorrente
                        (fornecedor_id, marca_concorrente, origem_cidade,
                         importada, importado_por, observacao, ativo)
                        VALUES (?,?,?,?,?,?,1)""",
                        (forn_sel[0], _mn, origem_n.strip() or None,
                         1 if importada else 0,
                         import_por.strip() or None,
                         obs_n.strip() or None))
                    conn.commit(); conn.close()
                    st.session_state.pop("show_form_marca", None)
                    st.success(f"✅ Marca '{_mn}' cadastrada!"); st.rerun()
        if cancelar:
            st.session_state.pop("show_form_marca", None); st.rerun()


def _form_editar_marca(cid, marca_at, origem_at, forns):
    # Carrega dados atuais do banco — 4 campos
    forn_ids = [f[0] for f in forns]
    fid_at   = query("""SELECT fornecedor_id, observacao, importada, importado_por
        FROM concorrente WHERE concorrente_id=?""", (cid,))
    if not fid_at:
        st.error("Marca não encontrada."); return

    idx_f          = forn_ids.index(fid_at[0][0]) if fid_at[0][0] in forn_ids else 0
    _importada_at  = bool(fid_at[0][2])
    _import_por_at = fid_at[0][3] or ""
    _obs_at        = fid_at[0][1] or ""

    # Form SEM condicional dentro — todos os campos sempre visíveis
    # Isso evita o bug do Streamlit onde st.form não reage a checkbox
    with st.container():
        with st.form(f"edit_marca_{cid}", clear_on_submit=False):
            col1, col2 = st.columns(2)
            with col1:
                forn_e  = st.selectbox("Fornecedor *", forns, index=idx_f,
                                       format_func=lambda x: x[1])
                marca_e = st.text_input("Nome *", value=marca_at)
                orig_e  = st.text_input("País / cidade de origem",
                                        value=origem_at or "",
                                        placeholder="Ex: França, Itália, São Paulo...")
            with col2:
                imp_e      = st.checkbox("🌍 Marca importada",
                                         value=_importada_at)
                import_e   = st.text_input("Importado por",
                                           value=_import_por_at,
                                           placeholder="Ex: La Pastina, BRF... (deixe vazio se não aplicável)")
                obs_e      = st.text_input("Observação", value=_obs_at)

            col_s, col_c = st.columns(2)
            with col_s: salvar   = st.form_submit_button("💾 Salvar", type="primary")
            with col_c: cancelar = st.form_submit_button("Cancelar")

        if salvar:
            if not marca_e.strip():
                st.error("Nome é obrigatório.")
            else:
                # Verifica duplicata (exceto o próprio registro)
                dup = query("""SELECT concorrente_id FROM concorrente
                    WHERE LOWER(marca_concorrente)=LOWER(?)
                      AND fornecedor_id=? AND ativo=1
                      AND concorrente_id != ?""",
                    (marca_e.strip(), forn_e[0], cid))
                if dup:
                    st.error(f"Já existe outra marca com o nome '{marca_e.strip()}' "
                             "para este fornecedor.")
                else:
                    conn = conectar()
                    conn.execute("""UPDATE concorrente SET
                        fornecedor_id=?, marca_concorrente=?, origem_cidade=?,
                        importada=?, importado_por=?, observacao=?
                        WHERE concorrente_id=?""",
                        (forn_e[0],
                         marca_e.strip(),
                         orig_e.strip() or None,
                         1 if imp_e else 0,
                         import_e.strip() or None,
                         obs_e.strip() or None,
                         cid))
                    conn.commit(); conn.close()
                    st.session_state.pop("marca_editar_id", None)
                    st.success("✅ Marca atualizada!"); st.rerun()

        if cancelar:
            st.session_state.pop("marca_editar_id", None); st.rerun()


def _confirmacao_excluir_marca(cid, marca):
    n_prods = query("SELECT COUNT(*) FROM produto_concorrente WHERE concorrente_id=?", (cid,))[0][0]
    n_pesq  = query("""SELECT COUNT(*) FROM pesquisa_preco_item pi
        JOIN produto_concorrente pc ON pi.produto_concorrente_id=pc.produto_concorrente_id
        WHERE pc.concorrente_id=?""", (cid,))[0][0]
    msg = f"Excluir **{marca}**?"
    if n_prods or n_pesq:
        msg += f" Remove {n_prods} produto(s) e {n_pesq} registro(s) de pesquisa."
    st.warning(msg)
    col1, col2 = st.columns(2)
    with col1:
        if st.button("Confirmar exclusão", key=f"conf_ex_{cid}", type="primary", width="stretch"):
            _excluir_marca(cid)
    with col2:
        if st.button("Cancelar", key=f"canc_ex_{cid}", width="stretch"):
            st.session_state.pop("marca_confirmar_excluir", None); st.rerun()


def _excluir_marca(cid):
    conn = conectar()
    pcs = [r[0] for r in conn.execute(
        "SELECT produto_concorrente_id FROM produto_concorrente WHERE concorrente_id=?", (cid,)).fetchall()]
    if pcs:
        ph = ",".join("?" * len(pcs))
        conn.execute(f"DELETE FROM produto_concorrente_relacao WHERE produto_concorrente_id IN ({ph})", pcs)
        conn.execute(f"DELETE FROM pesquisa_preco_item WHERE produto_concorrente_id IN ({ph})", pcs)
        conn.execute("DELETE FROM produto_concorrente WHERE concorrente_id=?", (cid,))
    conn.execute("DELETE FROM concorrente WHERE concorrente_id=?", (cid,))
    conn.commit(); conn.close()
    st.session_state.pop("marca_confirmar_excluir", None)
    st.success("Marca excluída!"); st.rerun()


# ==============================================================
# ABA 2 -- PRODUTOS E RELACOES (unificada)
# Ao cadastrar ou editar um produto, vincula ao produto proprio na mesma tela
# ==============================================================

def _produtos_e_relacoes():
    import io
    concs = query("""SELECT conc.concorrente_id, conc.marca_concorrente, f.nome_fantasia
        FROM concorrente conc JOIN fornecedor f ON conc.fornecedor_id=f.fornecedor_id
        WHERE conc.ativo=1 ORDER BY f.nome_fantasia, conc.marca_concorrente""")
    if not concs:
        st.info("Cadastre uma marca concorrente primeiro (aba Marcas concorrentes).")
        return

    cats  = cache_categorias()
    forns = cache_fornecedores()

    # ── Filtro de fornecedor — define o contexto do formulário ──────────────
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        forn_opts = [(None,"— Todos os fornecedores")] + [(f[0],f[1]) for f in forns]
        fil_forn  = st.selectbox("Fornecedor ref.", forn_opts,
                                 format_func=lambda x: x[1], key="pc_fil_forn")

    if st.button("➕ Novo produto concorrente", type="primary", key="btn_novo_pc"):
        st.session_state["show_form_pc"] = True
        st.session_state.pop("pc_editar_id", None)
        st.session_state.pop("pc_excluir_id", None)

    if st.session_state.get("show_form_pc"):
        # Passa o fornecedor selecionado no filtro para o form
        # Se nenhum selecionado, exige seleção antes de abrir o form
        if not fil_forn[0]:
            st.warning("⚠️ Selecione o **Fornecedor ref.** acima antes de cadastrar um produto concorrente.")
        else:
            _form_novo_produto_relacao(concs, cats, forns, fil_forn)

    st.divider()
    with col2:
        cats_disp = query("""
            SELECT DISTINCT cat.categoria_id, cat.nome_categoria
            FROM produto_concorrente pc
            JOIN concorrente conc ON pc.concorrente_id=conc.concorrente_id
            JOIN categoria cat    ON pc.categoria_id=cat.categoria_id
            WHERE cat.ativo=1 """ +
            ("AND conc.fornecedor_id=?" if fil_forn[0] else "") +
            " ORDER BY cat.nome_categoria",
            (fil_forn[0],) if fil_forn[0] else ()
        )
        cat_opts = [(None,"— Todas as categorias")] + [(c[0],c[1]) for c in cats_disp]
        fil_cat  = st.selectbox("Categoria", cat_opts,
                                format_func=lambda x: x[1], key="pc_fil_cat")
    with col3:
        fil_aud  = st.selectbox("Classificação",
                                ["Todos","📊 Auditáveis","🚫 Não auditáveis"],
                                key="pc_fil_aud")
    with col4:
        fil_busca = st.text_input("🔍 Buscar", placeholder="Nome ou marca...",
                                  key="pc_fil_busca")

    # ── Query com todos os filtros ────────────────────────────────────────
    where  = ["pc.ativo=1"]
    params = []
    if fil_forn[0]:
        where.append("conc.fornecedor_id=?"); params.append(fil_forn[0])
    if fil_cat[0]:
        where.append("pc.categoria_id=?");   params.append(fil_cat[0])
    if fil_aud == "📊 Auditáveis":
        where.append("COALESCE(pc.auditavel,1)=1")
    elif fil_aud == "🚫 Não auditáveis":
        where.append("COALESCE(pc.auditavel,1)=0")
    if fil_busca.strip():
        _term = fil_busca.strip().replace("%","%%").replace("_",r"\_")
        b = f"%{_term}%"
        where.append("(pc.descricao LIKE ? OR pc.descricao_curta LIKE ? OR conc.marca_concorrente LIKE ?)")
        params.extend([b, b, b])

    dados = query(f"""
        SELECT pc.produto_concorrente_id,
               conc.marca_concorrente,
               f.nome_fantasia,
               COALESCE(pc.descricao_curta, pc.descricao) AS desc_c,
               COALESCE(cat.nome_categoria,'—') AS categoria,
               pc.peso, pc.unidade_medida,
               COALESCE(pc.auditavel,1) AS auditavel,
               COUNT(rel.relacao_id) AS vinculos,
               COALESCE(pc.ean_concorrente,'') AS ean
        FROM produto_concorrente pc
        JOIN concorrente conc ON pc.concorrente_id=conc.concorrente_id
        JOIN fornecedor f     ON conc.fornecedor_id=f.fornecedor_id
        LEFT JOIN categoria cat ON pc.categoria_id=cat.categoria_id
        LEFT JOIN produto_concorrente_relacao rel
               ON rel.produto_concorrente_id=pc.produto_concorrente_id
        WHERE {' AND '.join(where)}
        GROUP BY pc.produto_concorrente_id, conc.marca_concorrente, f.nome_fantasia, pc.descricao_curta, pc.descricao, cat.nome_categoria, pc.peso, pc.unidade_medida, pc.auditavel, pc.ean_concorrente
        ORDER BY COALESCE(pc.auditavel,1) DESC,
                 conc.marca_concorrente, pc.descricao_curta
    """, tuple(params))

    if not dados:
        st.info("Nenhum produto encontrado para os filtros selecionados.")
        return

    # ── Cabeçalho com totais ──────────────────────────────────────────────
    total_todos = query("SELECT COUNT(*) FROM produto_concorrente WHERE ativo=1")[0][0]
    n_aud  = sum(1 for r in dados if r[7])
    n_naud = sum(1 for r in dados if not r[7])
    col_info, col_exp = st.columns([3,1])
    col_info.caption(
        f"**{len(dados)}** produto(s)" +
        (f" de {total_todos} total" if any([fil_forn[0],fil_cat[0],fil_busca.strip(),fil_aud!="Todos"]) else "") +
        f"  |  📊 {n_aud} auditável(is)  ·  🚫 {n_naud} não auditável(is)"
    )
    with col_exp:
        df_exp = pd.DataFrame([
            (r[1], r[2], r[4], r[3] or "—",
             f"{r[5]:.3f}".rstrip("0").rstrip(".") if r[5] else "—",
             r[6] or "—",
             "Auditável" if r[7] else "Não auditável",
             r[8], r[9] or "—")
            for r in dados
        ], columns=["Marca","Fornecedor ref.","Categoria","Produto",
                    "Peso","UM","Classificação","Vínculos","EAN"])
        buf = io.BytesIO()
        df_exp.to_excel(buf, index=False, sheet_name="Produtos concorrentes")
        buf.seek(0)
        st.download_button("⬇️ Exportar Excel", data=buf,
                           file_name="produtos_concorrentes.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           width="stretch")

    # ── Lista ─────────────────────────────────────────────────────────────
    hc = st.columns([1.4, 1.0, 1.0, 2.8, 0.7, 0.5, 1.0])
    for col, txt in zip(hc, ["Marca","Forn.","Categ.","Produto","Audit.","Vínculos",""]):
        col.markdown(f"<small><b>{txt}</b></small>", unsafe_allow_html=True)
    st.divider()

    for row in dados:
        pc_id, marca, forn_n, desc_c, cat_n, peso, um, auditavel, vinculos, ean = row

        c = st.columns([1.4, 1.0, 1.0, 2.8, 0.7, 0.5, 1.0])
        c[0].write(marca)
        c[1].caption(forn_n)
        c[2].caption(cat_n)
        # Produto com badges de EAN e não auditável
        _label = desc_c or "—"
        c[3].write(_label)
        if not auditavel:
            c[3].caption("🚫 Não auditável")
        elif ean:
            c[3].caption(f"EAN: {ean}")
        # Auditável clicável para alternar
        aud_ico = "📊" if auditavel else "🚫"
        if c[4].button(aud_ico, key=f"togaud_{pc_id}",
                       width="stretch",
                       help="Clique para alternar auditável/não auditável"):
            from database import conectar as _con
            conn = _con()
            conn.execute("UPDATE produto_concorrente SET auditavel=? WHERE produto_concorrente_id=?",
                         (0 if auditavel else 1, pc_id))
            conn.commit(); conn.close()
            st.rerun()
        c[5].caption(f"{vinculos}✅" if vinculos else "—")
        with c[6]:
            b1, b2 = st.columns(2)
            with b1:
                if st.button("✏️", key=f"edpc_{pc_id}", help="Editar",
                             width="stretch"):
                    st.session_state["pc_editar_id"] = pc_id
                    st.session_state.pop("pc_excluir_id", None)
                    st.session_state.pop("show_form_pc", None)
                    st.rerun()
            with b2:
                if st.button("🗑️", key=f"expc_{pc_id}", help="Excluir",
                             width="stretch"):
                    st.session_state["pc_excluir_id"] = pc_id
                    st.session_state.pop("pc_editar_id", None)
                    st.session_state.pop("show_form_pc", None)
                    st.rerun()

        if st.session_state.get("pc_editar_id") == pc_id:
            _form_editar_produto_relacao(pc_id, concs, cats, forns)
        if st.session_state.get("pc_excluir_id") == pc_id:
            _confirmacao_excluir_produto(pc_id, desc_c or f"#{pc_id}")


def _form_novo_produto_relacao(concs, cats, forns, forn_fil=None):
    """
    Cadastra produto concorrente.
    forn_fil = (forn_id, nome) do fornecedor selecionado no filtro.
    A classificação (auditável/não auditável) fica FORA do form
    para reagir imediatamente e mostrar/ocultar a seção de vínculo.
    """
    with st.container():
        st.markdown("**Novo produto concorrente**")

        # ── Classificação FORA do form — reage imediatamente ─────────────
        st.markdown("**1. Classificação**")
        _aud_key = "npc_auditavel"
        _eh_auditavel = st.radio(
            "Este produto:",
            [True, False],
            format_func=lambda x:
                "📊 Auditável — concorrente direto ou indireto (entra nas análises de preço e share)"
                if x else
                "🚫 Não auditável — está na mesma gôndola mas não concorre diretamente "
                "(registrado para contexto, ignorado nas métricas)",
            key=_aud_key,
            horizontal=False,
        )

        st.divider()
        st.markdown("**2. Dados do produto**")

        with st.form("novo_pc_rel", clear_on_submit=True):

            col1, col2 = st.columns(2)
            with col1:
                conc_sel = st.selectbox("Marca concorrente *", concs,
                                        format_func=lambda x: f"{x[1]}  ({x[2]})")
                desc   = st.text_input("Descrição completa *",
                                       placeholder="Ex: Vinagre de Maçã Castelo 750ml")
                desc_c = st.text_input("Descrição curta", max_chars=56,
                                       placeholder="Ex: Vin. Maçã Castelo 750ml")
                cat_sel = st.selectbox("Categoria",
                                       [(None,"— sem categoria —")] + list(cats),
                                       format_func=lambda x: x[1])
            with col2:
                peso  = st.number_input("Peso / volume", min_value=0.0, format="%.3f")
                um    = st.selectbox("Unidade", ["UN","kg","g","L","ml"])
                ean   = st.text_input("EAN-13 (opcional)",
                                      placeholder="Pode ser preenchido depois via Busca por EAN")
                obs_p = st.text_input("Observação")

            # ── Vínculo: só aparece para auditáveis ──────────────────────
            prod_ref = (None, "--", "--")
            tipo_rel = "direto"

            if _eh_auditavel:
                st.divider()
                # Usa fornecedor já selecionado no filtro — sem selectbox redundante
                forn_ref_id   = forn_fil[0]   if forn_fil else None
                forn_ref_nome = forn_fil[1]   if forn_fil else "—"
                st.markdown(f"**3. Vincular ao meu produto de referência — {forn_ref_nome}**")
                st.caption(
                    "Qual produto seu este concorrente disputa? "
                    "Pode editar ou adicionar vínculos depois.")
                if forn_ref_id:
                    prods_ref = query("""SELECT produto_id, codigo_produto, descricao_curta
                        FROM produto WHERE fornecedor_id=? AND ativo=1
                        ORDER BY descricao_curta""", (forn_ref_id,))
                    if prods_ref:
                        prod_ref = st.selectbox(
                            "Meu produto de referência",
                            [(None,"— Vincular depois —")] + list(prods_ref),
                            format_func=lambda x: x[1] if x[0] is None
                                        else f"{x[1]} — {x[2]}")
                        tipo_rel = st.selectbox(
                            "Tipo de relação", ["direto","indireto"],
                            help="Direto = mesmo produto. Indireto = substituto parcial.")
                    else:
                        st.caption("Nenhum produto cadastrado para este fornecedor. "
                                   "O vínculo pode ser feito depois.")

            col_s, col_c = st.columns(2)
            with col_s: salvar   = st.form_submit_button("💾 Salvar produto", type="primary")
            with col_c: cancelar = st.form_submit_button("Cancelar")

        if cancelar:
            st.session_state.pop("show_form_pc", None)
            st.session_state.pop(_aud_key, None)
            st.rerun()

        if salvar:
            if not desc.strip():
                st.error("Descrição completa é obrigatória."); return

            conn = conectar()
            conn.execute("""INSERT INTO produto_concorrente
                (concorrente_id, categoria_id, descricao, descricao_curta,
                 peso, unidade_medida, ean_concorrente, auditavel, observacao, ativo)
                VALUES (?,?,?,?,?,?,?,?,?,1)""",
                (conc_sel[0],
                 cat_sel[0] if cat_sel and cat_sel[0] else None,
                 desc.strip(), desc_c.strip() or None,
                 peso or None, um,
                 ean.strip() or None,
                 1 if _eh_auditavel else 0,
                 obs_p.strip() or None))
            conn.commit()

            pc_id_novo = conn.execute(
                "SELECT produto_concorrente_id FROM produto_concorrente "
                "ORDER BY produto_concorrente_id DESC LIMIT 1").fetchone()[0]

            if prod_ref and prod_ref[0]:
                try:
                    conn.execute("""INSERT OR IGNORE INTO produto_concorrente_relacao
                        (produto_id, produto_concorrente_id, tipo_relacao) VALUES (?,?,?)""",
                        (prod_ref[0], pc_id_novo, tipo_rel))
                    conn.commit()
                except Exception:
                    pass

            conn.close()
            st.session_state.pop("show_form_pc", None)
            st.session_state.pop(_aud_key, None)
            aud_label = "auditável" if _eh_auditavel else "não auditável"
            vinc_msg  = (f" · vinculado a '{prod_ref[2]}'"
                         if prod_ref and prod_ref[0] else "")
            st.success(
                f"✅ '{desc_c or desc}' cadastrado como {aud_label}{vinc_msg}!")
            st.rerun()


def _form_editar_produto_relacao(pc_id, concs, cats, forns):
    """Edita produto concorrente e gerencia seus vinculos na mesma tela."""
    prod = query("""SELECT pc.concorrente_id, pc.descricao, pc.descricao_curta,
               pc.categoria_id, pc.peso, pc.unidade_medida, pc.ean_concorrente, pc.observacao, pc.ativo
        FROM produto_concorrente pc WHERE pc.produto_concorrente_id=?""", (pc_id,))
    if not prod: return

    conc_id_at, desc_at, desc_c_at, cat_id_at, peso_at, um_at, ean_at, obs_at, ativo_at = prod[0]
    cats_lista = cache_categorias()
    ums = ["UN","kg","g","L","ml"]
    conc_ids  = [c[0] for c in concs]
    idx_conc  = conc_ids.index(conc_id_at) if conc_id_at in conc_ids else 0
    cats_opts = [(None,"-")] + list(cats_lista)
    cat_ids   = [c[0] for c in cats_opts]
    idx_cat   = cat_ids.index(cat_id_at) if cat_id_at in cat_ids else 0
    idx_um    = ums.index(um_at) if um_at in ums else 0

    with st.container():
        st.markdown(f"**Editando: {desc_c_at or desc_at}**")

        # -- Edicao dos dados do produto --
        with st.form(f"edit_pc_{pc_id}"):
            col1, col2 = st.columns(2)
            with col1:
                conc_e  = st.selectbox("Marca *", concs, index=idx_conc,
                                       format_func=lambda x: f"{x[1]} ({x[2]})", key=f"ep_c_{pc_id}")
                desc_e  = st.text_input("Descrição completa *", value=desc_at or "", key=f"ep_d_{pc_id}")
                desc_ce = st.text_input("Descrição curta", value=desc_c_at or "",
                                        max_chars=56, key=f"ep_dc_{pc_id}")
                cat_e   = st.selectbox("Categoria", cats_opts, index=idx_cat,
                                       format_func=lambda x: x[1], key=f"ep_cat_{pc_id}")
            with col2:
                peso_e  = st.number_input("Peso / volume", min_value=0.0,
                                          value=float(peso_at or 0), format="%.3f", key=f"ep_p_{pc_id}")
                um_e    = st.selectbox("Unidade", ums, index=idx_um, key=f"ep_u_{pc_id}")
                ean_e   = st.text_input("EAN", value=ean_at or "", key=f"ep_e_{pc_id}")
                obs_e   = st.text_input("Observação", value=obs_at or "", key=f"ep_o_{pc_id}")
                ativo_e = st.checkbox("Ativo", value=bool(ativo_at), key=f"ep_a_{pc_id}")
            col_s, col_c = st.columns(2)
            with col_s: salvar   = st.form_submit_button("Salvar produto", type="primary")
            with col_c: cancelar = st.form_submit_button("Cancelar")

        if salvar:
            if not desc_e.strip():
                st.error("Descrição é obrigatória.")
            else:
                conn = conectar()
                conn.execute("""UPDATE produto_concorrente SET
                    concorrente_id=?,descricao=?,descricao_curta=?,categoria_id=?,
                    peso=?,unidade_medida=?,ean_concorrente=?,observacao=?,ativo=?
                    WHERE produto_concorrente_id=?""",
                    (conc_e[0], desc_e.strip(), desc_ce.strip() or None,
                     cat_e[0] if cat_e and cat_e[0] else None,
                     peso_e or None, um_e, ean_e.strip() or None,
                     obs_e.strip() or None, 1 if ativo_e else 0, pc_id))
                conn.commit(); conn.close()
                st.session_state.pop("pc_editar_id", None)
                st.success("Produto atualizado!"); st.rerun()
        if cancelar:
            st.session_state.pop("pc_editar_id", None); st.rerun()

        # -- Vinculos existentes --
        st.divider()
        st.markdown("**Vinculos com meus produtos**")

        vinculos = query("""
            SELECT rel.relacao_id, p.codigo_produto, p.descricao_curta,
                   f.nome_fantasia, rel.tipo_relacao
            FROM produto_concorrente_relacao rel
            JOIN produto p      ON rel.produto_id=p.produto_id
            JOIN fornecedor f   ON p.fornecedor_id=f.fornecedor_id
            WHERE rel.produto_concorrente_id=?
            ORDER BY f.nome_fantasia, p.descricao_curta
        """, (pc_id,))

        if vinculos:
            st.caption(f"{len(vinculos)} vinculo(s) cadastrado(s)")
            for rel_id, cod, prod_n, forn_n, tipo_rel in vinculos:
                icone = "🎯 Direto" if tipo_rel == "direto" else "↔️ Indireto"
                c1, c2, c3, c4 = st.columns([2.5, 2.0, 1.5, 1.5])
                c1.write(f"{prod_n}")
                c2.caption(f"{forn_n} | {cod}")
                c3.caption(icone)
                with c4:
                    b1, b2 = st.columns(2)
                    with b1:
                        novo_tipo = "indireto" if tipo_rel == "direto" else "direto"
                        label_btn = "→ Indireto" if tipo_rel == "direto" else "→ Direto"
                        if st.button(label_btn, key=f"chg_rel_{rel_id}",
                                     width="stretch",
                                     help=f"Alterar para {novo_tipo}"):
                            conn = conectar()
                            conn.execute("""UPDATE produto_concorrente_relacao
                                SET tipo_relacao=? WHERE relacao_id=?""",
                                (novo_tipo, rel_id))
                            conn.commit(); conn.close()
                            st.success(f"Alterado para {novo_tipo}!")
                            st.rerun()
                    with b2:
                        if not st.session_state.get(f"conf_del_rel_{rel_id}"):
                            if st.button("🗑️", key=f"del_rel_{rel_id}",
                                         width="stretch",
                                         help="Remover vínculo"):
                                st.session_state[f"conf_del_rel_{rel_id}"] = True
                                st.rerun()
                        else:
                            st.warning("⚠️ Remover este vínculo?")
                            _r1, _r2 = st.columns(2)
                            with _r1:
                                if st.button("✅ Sim", key=f"conf_rel_ok_{rel_id}",
                                             type="primary", width="stretch"):
                                    conn = conectar()
                                    conn.execute("DELETE FROM produto_concorrente_relacao WHERE relacao_id=?", (rel_id,))
                                    conn.commit(); conn.close()
                                    st.session_state.pop(f"conf_del_rel_{rel_id}", None)
                                    st.rerun()
                            with _r2:
                                if st.button("❌ Não", key=f"conf_rel_no_{rel_id}",
                                             width="stretch"):
                                    st.session_state.pop(f"conf_del_rel_{rel_id}", None)
                                    st.rerun()
        else:
            st.caption("Nenhum vinculo cadastrado para este produto.")

        # -- Adicionar novo vinculo --
        st.divider()
        st.markdown("**Adicionar vinculo**")

        ids_ja = {r[0] for r in query(
            "SELECT produto_id FROM produto_concorrente_relacao WHERE produto_concorrente_id=?",
            (pc_id,))}

        forn_add = st.selectbox("Fornecedor", forns, format_func=lambda x: x[1],
                                key=f"fadd_{pc_id}")
        prods_disp = [p for p in query("""SELECT produto_id, codigo_produto, descricao_curta
            FROM produto WHERE fornecedor_id=? AND ativo=1 ORDER BY descricao_curta""",
            (forn_add[0],)) if p[0] not in ids_ja]

        if not prods_disp:
            st.caption("Todos os produtos deste fornecedor já estão vinculados.")
        else:
            with st.form(f"add_rel_{pc_id}", clear_on_submit=True):
                prod_add = st.selectbox("Meu produto a vincular", prods_disp,
                                        format_func=lambda x: f"{x[1]} — {x[2]}")
                tipo_add = st.selectbox("Tipo de relação", ["direto","indireto"])
                if st.form_submit_button("Vincular", type="primary"):
                    conn = conectar()
                    try:
                        conn.execute("""INSERT OR IGNORE INTO produto_concorrente_relacao
                            (produto_id, produto_concorrente_id, tipo_relacao) VALUES (?,?,?)""",
                            (prod_add[0], pc_id, tipo_add))
                        conn.commit()
                        st.success(f"'{prod_add[2]}' vinculado!")
                        st.rerun()
                    except Exception as e:
                        st.error(str(e))
                    finally:
                        conn.close()


def _confirmacao_excluir_produto(pc_id, desc):
    n_rel  = query("SELECT COUNT(*) FROM produto_concorrente_relacao WHERE produto_concorrente_id=?", (pc_id,))[0][0]
    n_pesq = query("SELECT COUNT(*) FROM pesquisa_preco_item WHERE produto_concorrente_id=?", (pc_id,))[0][0]
    msg = f"Excluir **{desc}**?"
    if n_rel or n_pesq:
        msg += f" Remove {n_rel} vínculo(s) e {n_pesq} registro(s) de pesquisa."
    st.warning(msg)
    col1, col2 = st.columns(2)
    with col1:
        if st.button("Confirmar exclusão", key=f"conf_expc_{pc_id}",
                     type="primary", width="stretch"):
            conn = conectar()
            conn.execute("DELETE FROM produto_concorrente_relacao WHERE produto_concorrente_id=?", (pc_id,))
            conn.execute("DELETE FROM pesquisa_preco_item WHERE produto_concorrente_id=?", (pc_id,))
            conn.execute("DELETE FROM produto_concorrente WHERE produto_concorrente_id=?", (pc_id,))
            conn.commit(); conn.close()
            st.session_state.pop("pc_excluir_id", None)
            st.success("Excluído!"); st.rerun()
    with col2:
        if st.button("Cancelar", key=f"canc_expc_{pc_id}", width="stretch"):
            st.session_state.pop("pc_excluir_id", None); st.rerun()


# ==============================================================
# ABA 3 -- BUSCA E GESTÃO POR EAN
# ==============================================================

def _gestao_por_ean():
    """Busca, cadastra e classifica concorrentes via EAN."""
    import io
    st.subheader("🔍 Gestão de concorrentes por EAN")
    st.caption(
        "Digite o EAN-13 de um produto encontrado na gôndola. "
        "O sistema verifica se já está cadastrado, se é auditável, "
        "e permite cadastrar ou reclassificar rapidamente."
    )

    forns = cache_fornecedores()
    if not forns:
        st.info("Cadastre um fornecedor primeiro."); return

    col1, col2 = st.columns([2,3])
    with col1:
        forn_sel = st.selectbox("Fornecedor de referência", forns,
                                format_func=lambda x: x[1], key="ean_forn")
    with col2:
        ean_busca = st.text_input("EAN-13",
                                  placeholder="Ex: 7891234567890",
                                  key="ean_busca",
                                  help="Cole ou digite o código de barras do produto")

    if ean_busca.strip():
        ean = ean_busca.strip()

        # ── Verifica se é produto NOSSO ──────────────────────────────────
        prod_nosso = query("""
            SELECT p.produto_id, p.descricao_curta, p.descricao,
                   f.nome_fantasia
            FROM produto p
            JOIN fornecedor f ON p.fornecedor_id=f.fornecedor_id
            WHERE p.ean=? AND p.ativo=1""", (ean,))

        if prod_nosso:
            pid, desc_c, desc, forn_n = prod_nosso[0]
            st.success(f"✅ **Produto próprio** — {forn_n}: {desc_c or desc}")
            st.info("Este EAN pertence a um produto da sua representação. Não é concorrente.")
            return

        # ── Verifica se já está cadastrado como concorrente ───────────────
        conc_existente = query("""
            SELECT pc.produto_concorrente_id, pc.descricao_curta, pc.descricao,
                   conc.marca_concorrente, pc.auditavel,
                   pc.ean_concorrente
            FROM produto_concorrente pc
            JOIN concorrente conc ON pc.concorrente_id=conc.concorrente_id
            WHERE pc.ean_concorrente=? AND pc.ativo=1""", (ean,))

        if conc_existente:
            pcid, desc_c, desc, marca, auditavel, ean_c = conc_existente[0]
            status_aud = "📊 Auditável" if auditavel else "🚫 Não auditável"
            st.info(
                f"**Já cadastrado:** {marca} — {desc_c or desc}  |  {status_aud}")

            # Permite reclassificar
            col_a, col_b = st.columns(2)
            if auditavel:
                if col_a.button("🚫 Marcar como não auditável", key="ean_naud",
                                width="stretch"):
                    conn = query.__self__ if hasattr(query,'__self__') else None
                    from database import conectar as _con
                    conn = _con()
                    conn.execute("UPDATE produto_concorrente SET auditavel=0 WHERE produto_concorrente_id=?",
                                 (pcid,))
                    conn.commit(); conn.close()
                    st.success("Marcado como não auditável — não entrará nas análises.")
                    st.rerun()
            else:
                if col_a.button("📊 Marcar como auditável", key="ean_aud",
                                width="stretch"):
                    from database import conectar as _con
                    conn = _con()
                    conn.execute("UPDATE produto_concorrente SET auditavel=1 WHERE produto_concorrente_id=?",
                                 (pcid,))
                    conn.commit(); conn.close()
                    st.success("Marcado como auditável — entrará nas análises.")
                    st.rerun()

            # Produtos nossos vinculados
            rels = query("""SELECT p.descricao_curta, rel.tipo_relacao
                FROM produto_concorrente_relacao rel
                JOIN produto p ON rel.produto_id=p.produto_id
                WHERE rel.produto_concorrente_id=?""", (pcid,))
            if rels:
                st.caption("Vinculado aos nossos produtos: " +
                           ", ".join(f"{r[0]} ({r[1]})" for r in rels))
            return

        # ── EAN desconhecido — primeiro verifica produtos sem EAN ──────────
        # Oferta: vincular EAN a produto já cadastrado sem EAN
        sem_ean = query("""
            SELECT pc.produto_concorrente_id, pc.descricao_curta, pc.descricao,
                   conc.marca_concorrente
            FROM produto_concorrente pc
            JOIN concorrente conc ON pc.concorrente_id=conc.concorrente_id
            WHERE conc.fornecedor_id=?
              AND (pc.ean_concorrente IS NULL OR pc.ean_concorrente='')
              AND pc.ativo=1
            ORDER BY conc.marca_concorrente, pc.descricao_curta""", (forn_sel[0],))

        st.warning(f"⚠️ EAN **{ean}** não encontrado na base.")
        st.divider()

        if sem_ean:
            with st.expander(f"🔗 Vincular a produto já cadastrado sem EAN ({len(sem_ean)} disponíveis)"):
                st.caption(
                    "Se este EAN pertence a um produto já cadastrado que estava "
                    "sem código de barras, selecione abaixo para vincular sem precisar recadastrar.")
                opts_sem = [(None,"— não é nenhum destes —")] +                            [(s[0], f"{s[3]} — {s[1] or s[2]}") for s in sem_ean]
                sel_sem = st.selectbox("Selecione o produto",
                                       opts_sem,
                                       format_func=lambda x: x[1],
                                       key="ean_vinc_sem")
                if sel_sem and sel_sem[0]:
                    st.info(f"Vai vincular EAN **{ean}** a: **{sel_sem[1]}**")
                    if st.button("✅ Confirmar vinculação", key="ean_vinc_ok",
                                 type="primary", width="stretch"):
                        from database import conectar as _con
                        conn = _con()
                        conn.execute(
                            "UPDATE produto_concorrente SET ean_concorrente=? WHERE produto_concorrente_id=?",
                            (ean, sel_sem[0]))
                        conn.commit(); conn.close()
                        st.success(f"✅ EAN {ean} vinculado com sucesso!")
                        st.session_state.pop("ean_busca", None)
                        st.rerun()

        # Produto com EAN diferente — oferta de correção
        conc_ean_dif = query("""
            SELECT pc.produto_concorrente_id, pc.descricao_curta,
                   pc.ean_concorrente, conc.marca_concorrente
            FROM produto_concorrente pc
            JOIN concorrente conc ON pc.concorrente_id=conc.concorrente_id
            WHERE conc.fornecedor_id=?
              AND pc.ean_concorrente IS NOT NULL AND pc.ean_concorrente != ''
              AND pc.ativo=1
            ORDER BY conc.marca_concorrente, pc.descricao_curta""", (forn_sel[0],))

        if conc_ean_dif:
            with st.expander("🔧 Corrigir EAN de produto já cadastrado"):
                st.caption(
                    "Se este é o EAN correto de um produto que foi cadastrado com "
                    "código errado, selecione e confirme a correção.")
                opts_cor = [(None,"— não é correção —")] +                            [(s[0], f"{s[3]} — {s[1]} (atual: {s[2]})") for s in conc_ean_dif]
                sel_cor = st.selectbox("Produto com EAN errado",
                                       opts_cor,
                                       format_func=lambda x: x[1],
                                       key="ean_cor_sel")
                if sel_cor and sel_cor[0]:
                    prod_info = next(s for s in conc_ean_dif if s[0]==sel_cor[0])
                    st.warning(
                        f"⚠️ EAN atual: **{prod_info[2]}**  →  Novo EAN: **{ean}**  "
                        f"| Produto: {prod_info[1]}")
                    if st.button("✅ Confirmar correção de EAN", key="ean_cor_ok",
                                 type="primary", width="stretch"):
                        from database import conectar as _con
                        conn = _con()
                        conn.execute(
                            "UPDATE produto_concorrente SET ean_concorrente=? WHERE produto_concorrente_id=?",
                            (ean, sel_cor[0]))
                        conn.commit(); conn.close()
                        st.success(f"✅ EAN corrigido: {prod_info[2]} → {ean}")
                        st.session_state.pop("ean_busca", None)
                        st.rerun()

        st.markdown("**Ou cadastrar como produto novo:**")
        st.divider()

        cats = cache_categorias()
        marcas_existing = query("""SELECT conc.concorrente_id, conc.marca_concorrente
            FROM concorrente conc WHERE conc.fornecedor_id=? AND conc.ativo=1
            ORDER BY conc.marca_concorrente""", (forn_sel[0],))

        col_a, col_b = st.columns(2)
        with col_a:
            marca_opts = ["➕ Nova marca..."] + [m[1] for m in marcas_existing]
            marca_sel  = st.selectbox("Marca", marca_opts, key="ean_marca")
            if marca_sel == "➕ Nova marca...":
                nova_marca = st.text_input("Nome da nova marca *",
                                           key="ean_nova_marca",
                                           placeholder="Ex: Castelo, Heinz...")
            else:
                nova_marca = ""

            desc_ean  = st.text_input("Descrição completa *",
                                      key="ean_desc",
                                      placeholder="Ex: Vinagre de Álcool 750ml")
            desc_c_ean = st.text_input("Descrição curta", max_chars=56,
                                       key="ean_desc_c")

        with col_b:
            cat_sel   = st.selectbox("Categoria", [(None,"— sem categoria —")] + list(cats),
                                     format_func=lambda x: x[1], key="ean_cat")
            peso_ean  = st.number_input("Peso/volume", min_value=0.0,
                                        format="%.3f", key="ean_peso")
            um_ean    = st.selectbox("Unidade", ["UN","kg","g","L","ml"], key="ean_um")

            auditavel_ean = st.radio(
                "Classificação",
                ["📊 Auditável — concorrente direto ou indireto",
                 "🚫 Não auditável — mesma gôndola, não concorre"],
                key="ean_class",
                help="Auditável = entra nas análises de preço e share. "
                     "Não auditável = registrado para contexto mas ignorado nas métricas.")

            # Produto nosso vinculado
            prods_nossos = query("""SELECT produto_id, descricao_curta, codigo_produto
                FROM produto WHERE fornecedor_id=? AND ativo=1
                ORDER BY descricao_curta""", (forn_sel[0],))
            prod_opts = [(None,"— sem vínculo —")] +                         [(p[0], f"{p[1]} ({p[2]})") for p in prods_nossos]
            prod_vin  = st.selectbox(
                "Vincular ao nosso produto (opcional para não auditáveis)",
                prod_opts,
                format_func=lambda x: x[1],
                key="ean_prod_vin",
                help="Obrigatório para auditáveis — identifica qual produto nosso este concorrente disputa. "
                     "Pode ser deixado em branco para produtos não auditáveis.")
            tipo_rel_ean = "direto"
            if prod_vin and prod_vin[0]:
                tipo_rel_ean = st.selectbox("Tipo de relação",
                                            ["direto","indireto"],
                                            key="ean_tipo_rel")

        if st.button("💾 Cadastrar concorrente", type="primary",
                     width="stretch", key="ean_salvar"):
            _nome = nova_marca.strip() if marca_sel=="➕ Nova marca..." else marca_sel
            _desc = st.session_state.get("ean_desc","").strip()
            if not _desc:
                st.error("Descrição é obrigatória."); return
            if marca_sel=="➕ Nova marca..." and not _nome:
                st.error("Informe o nome da nova marca."); return

            from database import conectar as _con, query as _q
            conn = _con()
            # Marca
            if marca_sel == "➕ Nova marca...":
                _dup = _q("SELECT concorrente_id FROM concorrente WHERE LOWER(marca_concorrente)=LOWER(?) AND fornecedor_id=? AND ativo=1",
                          (_nome, forn_sel[0]))
                if _dup:
                    st.error(f"⚠️ A marca **{_nome}** já está cadastrada para este fornecedor.")
                    conn.close(); return
                cur = conn.cursor()
                cur.execute("INSERT INTO concorrente (fornecedor_id, marca_concorrente, ativo) VALUES (?,?,1)",
                            (forn_sel[0], _nome))
                conc_id = cur.lastrowid
            else:
                idx_m = [m[1] for m in marcas_existing].index(marca_sel)
                conc_id = marcas_existing[idx_m][0]

            # Produto concorrente
            _aud = 0 if "Não auditável" in auditavel_ean else 1
            cur  = conn.cursor()
            cur.execute("""INSERT INTO produto_concorrente
                (concorrente_id, categoria_id, descricao, descricao_curta,
                 peso, unidade_medida, ean_concorrente, auditavel, ativo)
                VALUES (?,?,?,?,?,?,?,?,1)""",
                (conc_id,
                 st.session_state.get("ean_cat",(None,))[0] if isinstance(
                     st.session_state.get("ean_cat"),(tuple,list)) else None,
                 _desc,
                 st.session_state.get("ean_desc_c","").strip() or None,
                 st.session_state.get("ean_peso",0) or None,
                 st.session_state.get("ean_um","UN"),
                 ean, _aud))
            pc_novo = cur.lastrowid

            # Vínculo com produto nosso
            _pv = st.session_state.get("ean_prod_vin")
            if _pv and isinstance(_pv,(tuple,list)) and _pv[0]:
                conn.execute("""INSERT OR IGNORE INTO produto_concorrente_relacao
                    (produto_id, produto_concorrente_id, tipo_relacao) VALUES (?,?,?)""",
                    (_pv[0], pc_novo, tipo_rel_ean))

            conn.commit(); conn.close()

            aud_label = "auditável" if _aud else "não auditável"
            st.success(f"✅ **{_desc}** cadastrado como {aud_label}!")
            for k in ["ean_busca","ean_desc","ean_desc_c","ean_nova_marca",
                      "ean_peso"]:
                st.session_state.pop(k, None)
            st.rerun()

    st.divider()

    # ── Lista de todos os concorrentes com classificação ─────────────────
    st.markdown("**Todos os concorrentes cadastrados**")
    col_f1, col_f2, col_f3 = st.columns(3)
    with col_f1:
        forn_lista = st.selectbox("Fornecedor", forns,
                                  format_func=lambda x: x[1], key="ean_lista_forn")
    with col_f2:
        classif_fil = st.selectbox("Classificação",
                                   ["Todos","📊 Auditáveis","🚫 Não auditáveis"],
                                   key="ean_lista_class")
    with col_f3:
        busca_ean_lista = st.text_input("🔍 Buscar", key="ean_lista_busca",
                                        placeholder="Nome, marca ou EAN...")

    where_l = ["conc.fornecedor_id=?"]
    params_l = [forn_lista[0]]
    if classif_fil == "📊 Auditáveis":
        where_l.append("COALESCE(pc.auditavel,1)=1")
    elif classif_fil == "🚫 Não auditáveis":
        where_l.append("COALESCE(pc.auditavel,1)=0")
    if busca_ean_lista.strip():
        b = f"%{busca_ean_lista.strip()}%"
        where_l.append("(pc.descricao LIKE ? OR pc.descricao_curta LIKE ? OR conc.marca_concorrente LIKE ? OR pc.ean_concorrente LIKE ?)")
        params_l.extend([b, b, b, b])

    lista = query(f"""
        SELECT pc.produto_concorrente_id,
               conc.marca_concorrente,
               pc.descricao_curta, pc.descricao,
               COALESCE(pc.ean_concorrente,'—'),
               COALESCE(cat.nome_categoria,'—'),
               COALESCE(pc.auditavel,1)
        FROM produto_concorrente pc
        JOIN concorrente conc ON pc.concorrente_id=conc.concorrente_id
        LEFT JOIN categoria cat ON pc.categoria_id=cat.categoria_id
        WHERE {' AND '.join(where_l)} AND pc.ativo=1
        ORDER BY conc.marca_concorrente, pc.descricao_curta
    """, tuple(params_l))

    if not lista:
        st.info("Nenhum concorrente encontrado.")
        return

    st.caption(f"{len(lista)} produto(s)")
    hc = st.columns([1.5, 2.5, 1.5, 1.5, 1])
    for col, txt in zip(hc, ["Marca","Produto","EAN","Categoria","Audit."]):
        col.markdown(f"<small><b>{txt}</b></small>", unsafe_allow_html=True)
    st.divider()

    for row in lista:
        pcid, marca, desc_c, desc, ean_c, cat_n, aud = row
        c = st.columns([1.5, 2.5, 1.5, 1.5, 1])
        c[0].caption(marca)
        c[1].write(desc_c or desc)
        c[2].caption(ean_c)
        c[3].caption(cat_n)
        aud_ico = "📊" if aud else "🚫"
        if c[4].button(aud_ico, key=f"ean_tog_{pcid}",
                       width="stretch",
                       help="Clique para alternar auditável/não auditável"):
            from database import conectar as _con
            conn = _con()
            conn.execute("UPDATE produto_concorrente SET auditavel=? WHERE produto_concorrente_id=?",
                         (0 if aud else 1, pcid))
            conn.commit(); conn.close()
            st.rerun()


# ==============================================================
# ABA 4 -- IMPORTAÇÃO EM MASSA DE CONCORRENTES VIA EXCEL
# ==============================================================

def _importar_concorrentes():
    """Importação em massa de produtos concorrentes via planilha Excel."""
    import io
    import pandas as pd

    st.subheader("📥 Importar concorrentes via Excel")
    st.caption(
        "Importe uma lista completa de produtos concorrentes. "
        "O sistema criará marcas novas automaticamente, vinculará aos produtos "
        "Specialli pelo código informado e classificará cada produto como "
        "auditável ou não auditável."
    )

    # ── Download do template ─────────────────────────────────────────────
    with st.expander("📄 Baixar template de importação", expanded=True):
        st.caption(
            "Use este template para preencher seus dados. "
            "Mantenha os nomes das colunas exatamente como estão. "
            "A coluna **produto_specialli_vinculado** aceita o código do produto "
            "(ex: 7, 11, 17, 1580) ou pode ficar vazia."
        )
        buf_tpl = _gerar_template_importacao()
        st.download_button(
            "⬇️ Baixar template Excel",
            data=buf_tpl,
            file_name="template_importacao_concorrentes.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch"
        )

    st.divider()

    # ── Upload do arquivo ─────────────────────────────────────────────────
    forns = cache_fornecedores()
    if not forns:
        st.info("Cadastre um fornecedor primeiro."); return

    # ── AVISO EXPLÍCITO — seleção de fornecedor obrigatória ─────────────────
    st.warning(
        "⚠️ **Selecione corretamente o fornecedor** antes de importar. "
        "Todos os produtos e marcas serão vinculados ao fornecedor escolhido. "
        "Importar com fornecedor errado gera dados incorretos difíceis de corrigir."
    )

    col1, col2 = st.columns(2)
    with col1:
        # Opções com placeholder explícito para forçar seleção consciente
        forn_opts = [(None, "— Selecione o fornecedor —")] + list(forns)
        forn_sel_raw = st.selectbox(
            "Fornecedor de referência *",
            forn_opts,
            format_func=lambda x: x[1],
            key="imp_conc_forn",
            help="Fornecedor cujos produtos concorrentes você está importando")
        forn_sel = forn_sel_raw if forn_sel_raw[0] else None

    with col2:
        arquivo = st.file_uploader(
            "Selecione o arquivo Excel (.xlsx)",
            type=["xlsx"],
            key="imp_conc_file"
        )

    if not forn_sel:
        st.info("☝️ Selecione o fornecedor de referência para continuar.")
        return

    if not arquivo:
        st.info("⬆️ Selecione o fornecedor acima e faça o upload do arquivo Excel.")
        return

    # Confirmação visual do fornecedor selecionado
    st.success(f"✅ Importando concorrentes para: **{forn_sel[1]}**")

    # ── Leitura e validação ───────────────────────────────────────────────
    try:
        df = pd.read_excel(arquivo, dtype=str)
        df.columns = [c.strip().lower().replace(" ","_") for c in df.columns]
    except Exception as e:
        st.error(f"Erro ao ler o arquivo: {e}"); return

    # Remove linhas de instrução (começam com #)
    if "marca" in df.columns:
        df = df[~df["marca"].fillna("").str.startswith("#")]
    df = df.dropna(subset=["marca","descricao"])
    df = df.reset_index(drop=True)

    if df.empty:
        st.error("Nenhum dado encontrado no arquivo. Verifique se preencheu corretamente.")
        return

    # Colunas obrigatórias
    colunas_obrig = ["marca","descricao"]
    faltando = [c for c in colunas_obrig if c not in df.columns]
    if faltando:
        st.error(f"Colunas obrigatórias não encontradas: {', '.join(faltando)}")
        st.caption(f"Colunas encontradas no arquivo: {', '.join(df.columns.tolist())}")
        return

    # Preenche colunas opcionais ausentes
    for col_op, default in [
        ("descricao_curta",""), ("peso",""), ("unidade_medida","g"),
        ("ean",""), ("categoria",""), ("auditavel","1"),
        ("tipo_relacao","indireto"), ("produto_specialli_vinculado",""),
        ("observacao","")
    ]:
        if col_op not in df.columns:
            df[col_op] = default

    # ── Preview dos dados ─────────────────────────────────────────────────
    st.success(f"✅ **{len(df)} produto(s)** encontrado(s) no arquivo.")

    n_aud  = (df["auditavel"].fillna("1").astype(str).str.strip() == "1").sum()
    n_naud = len(df) - n_aud
    n_direto   = (df["tipo_relacao"].fillna("").str.strip() == "direto").sum()
    n_indireto = (df["tipo_relacao"].fillna("").str.strip() == "indireto").sum()
    marcas_unicas = df["marca"].nunique()

    col_a, col_b, col_c, col_d = st.columns(4)
    col_a.metric("Marcas únicas", marcas_unicas)
    col_b.metric("📊 Auditáveis", n_aud)
    col_c.metric("🚫 Não auditáveis", n_naud)
    col_d.metric("Diretos / Indiretos", f"{n_direto} / {n_indireto}")

    # Preview tabela
    with st.expander("👁 Visualizar dados antes de importar", expanded=False):
        cols_show = ["marca","descricao_curta","peso","unidade_medida",
                     "auditavel","tipo_relacao","produto_specialli_vinculado"]
        cols_show = [c for c in cols_show if c in df.columns]
        st.dataframe(df[cols_show].head(30), width="stretch",
                     hide_index=True)
        if len(df) > 30:
            st.caption(f"Mostrando 30 de {len(df)} linhas.")

    # ── Verificação prévia de duplicatas ──────────────────────────────────
    existentes = query("""
        SELECT pc.descricao, pc.descricao_curta, conc.marca_concorrente
        FROM produto_concorrente pc
        JOIN concorrente conc ON pc.concorrente_id=conc.concorrente_id
        WHERE conc.fornecedor_id=? AND pc.ativo=1
    """, (forn_sel[0],))
    existentes_set = {
        (r[2].strip().lower(), (r[1] or r[0] or "").strip().lower())
        for r in existentes
    }

    duplicatas = []
    novos = []
    for _, row in df.iterrows():
        marca_k = str(row.get("marca","")).strip().lower()
        desc_c  = str(row.get("descricao_curta","")).strip().lower()
        desc    = str(row.get("descricao","")).strip().lower()
        chave   = (marca_k, desc_c if desc_c else desc)
        if chave in existentes_set:
            duplicatas.append(row.get("descricao",""))
        else:
            novos.append(row)

    if duplicatas:
        st.warning(
            f"⚠️ **{len(duplicatas)} produto(s) já cadastrado(s)** "
            "serão ignorados na importação:"
        )
        for d in duplicatas[:5]:
            st.caption(f"  • {d}")
        if len(duplicatas) > 5:
            st.caption(f"  ... e mais {len(duplicatas)-5}")

    n_importar = len(novos)
    if n_importar == 0:
        st.info("Todos os produtos já estão cadastrados. Nada a importar.")
        return

    st.info(f"**{n_importar} produto(s) novo(s)** serão importados.")

    # ── Opções de importação ──────────────────────────────────────────────
    col_op1, col_op2 = st.columns(2)
    with col_op1:
        ignorar_sem_vinculo = st.checkbox(
            "Pular produtos sem vínculo com produto Specialli",
            value=False,
            key="imp_skip_sem_vinculo",
            help="Se marcado, produtos sem código na coluna produto_specialli_vinculado "
                 "só serão importados se forem não auditáveis")
    with col_op2:
        criar_marcas_novas = st.checkbox(
            "Criar marcas novas automaticamente",
            value=True,
            key="imp_criar_marcas",
            help="Se desmarcado, produtos de marcas não cadastradas serão pulados")

    # ── Botão de importação ───────────────────────────────────────────────
    st.divider()
    if st.button(f"📥 Importar {n_importar} produto(s)",
                 type="primary", width="stretch",
                 key="btn_importar_conc"):

        from database import conectar as _con
        conn = _con()

        # Cache de marcas existentes
        marcas_db = {}
        for conc in query("""SELECT concorrente_id, LOWER(marca_concorrente)
            FROM concorrente WHERE fornecedor_id=? AND ativo=1""",
            (forn_sel[0],)):
            marcas_db[conc[1]] = conc[0]

        # Cache de produtos Specialli por código
        prods_sp = {}
        for p in query("""SELECT produto_id, codigo_produto
            FROM produto WHERE fornecedor_id=? AND ativo=1""",
            (forn_sel[0],)):
            prods_sp[str(p[1]).strip().lower()] = p[0]

        # Cache de categorias
        cats_db = {}
        for c in query("SELECT categoria_id, LOWER(nome_categoria) FROM categoria WHERE ativo=1"):
            cats_db[c[1]] = c[0]

        importados = 0
        pulados    = 0
        erros      = []

        for _, row in pd.DataFrame(novos).iterrows():
            try:
                marca_nm = str(row.get("marca","")).strip()
                if not marca_nm:
                    pulados += 1; continue

                desc      = str(row.get("descricao","")).strip()
                desc_c    = str(row.get("descricao_curta","")).strip() or None
                ean       = str(row.get("ean","")).strip() or None
                obs       = str(row.get("observacao","")).strip() or None
                categoria = str(row.get("categoria","")).strip()
                tipo_rel  = str(row.get("tipo_relacao","indireto")).strip().lower()
                if tipo_rel not in ("direto","indireto"):
                    tipo_rel = "indireto"

                # auditavel
                aud_raw = str(row.get("auditavel","1")).strip()
                auditavel = 1 if aud_raw in ("1","sim","yes","true","s") else 0

                # peso
                try:
                    peso = float(str(row.get("peso","")).replace(",",".")) or None
                except:
                    peso = None

                um = str(row.get("unidade_medida","g")).strip() or "g"

                # cod vinculado
                cod_vinc = str(row.get("produto_specialli_vinculado","")).strip().lower()
                prod_id_vinc = prods_sp.get(cod_vinc) if cod_vinc else None

                # Regra: auditável sem vínculo pode ser pulado
                if ignorar_sem_vinculo and auditavel == 1 and not prod_id_vinc:
                    pulados += 1; continue

                # Categoria
                cat_id = cats_db.get(categoria.lower()) if categoria else None

                # Marca — cria se não existir
                marca_key = marca_nm.lower()
                if marca_key not in marcas_db:
                    if not criar_marcas_novas:
                        pulados += 1; continue
                    # Verifica duplicata
                    dup = conn.execute(
                        "SELECT concorrente_id FROM concorrente "
                        "WHERE LOWER(marca_concorrente)=? AND fornecedor_id=? AND ativo=1",
                        (marca_key, forn_sel[0])).fetchone()
                    if dup:
                        marcas_db[marca_key] = dup[0]
                    else:
                        cur = conn.cursor()
                        cur.execute(
                            "INSERT INTO concorrente (fornecedor_id, marca_concorrente, ativo) VALUES (?,?,1)",
                            (forn_sel[0], marca_nm))
                        marcas_db[marca_key] = cur.lastrowid
                        conn.commit()

                conc_id = marcas_db[marca_key]

                # Produto concorrente
                cur = conn.cursor()
                cur.execute("""INSERT INTO produto_concorrente
                    (concorrente_id, categoria_id, descricao, descricao_curta,
                     peso, unidade_medida, ean_concorrente, auditavel,
                     observacao, ativo)
                    VALUES (?,?,?,?,?,?,?,?,?,1)""",
                    (conc_id, cat_id, desc, desc_c,
                     peso, um, ean, auditavel, obs))
                pc_id = cur.lastrowid
                conn.commit()

                # Vínculo com produto Specialli
                if prod_id_vinc:
                    conn.execute("""INSERT OR IGNORE INTO produto_concorrente_relacao
                        (produto_id, produto_concorrente_id, tipo_relacao)
                        VALUES (?,?,?)""",
                        (prod_id_vinc, pc_id, tipo_rel))
                    conn.commit()

                importados += 1

            except Exception as e:
                erros.append(f"{row.get('descricao','?')}: {e}")

        conn.close()

        # Resultado
        if importados > 0:
            st.success(
                f"✅ **{importados} produto(s) importado(s)** com sucesso!\n"
                f"{'🔗 Vínculos criados para produtos com código Specialli.' if prod_id_vinc else ''}")
        if pulados > 0:
            st.info(f"ℹ️ {pulados} produto(s) pulados (sem vínculo ou marca não criada).")
        if erros:
            st.error(f"❌ {len(erros)} erro(s):")
            for e in erros[:5]:
                st.caption(f"  • {e}")

        if importados > 0:
            st.balloons()
            st.session_state["cc_aba"] = "prods"
            st.rerun()


def _gerar_template_importacao():
    """Gera o arquivo Excel template para importação de concorrentes."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Concorrentes_Importacao"

    COR_H = "2D6A4F"
    COR_I = "F5F5F5"
    borda = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin'))

    headers = [
        ("marca",                        "Marca *",                       20),
        ("descricao",                     "Descrição Completa *",          50),
        ("descricao_curta",               "Descrição Curta (máx 56)",      35),
        ("peso",                          "Peso (número)",                  10),
        ("unidade_medida",                "UM (g / Kg / UN)",               12),
        ("ean",                           "EAN-13",                         18),
        ("categoria",                     "Categoria",                      28),
        ("auditavel",                     "Auditável (1=sim / 0=não) *",    14),
        ("tipo_relacao",                  "Tipo Relação (direto/indireto)*",16),
        ("produto_specialli_vinculado",   "Código Produto Specialli",       22),
        ("observacao",                    "Observação",                     35),
    ]

    # Linha 1 — chave do campo (usada pelo sistema)
    for col, (chave, _, largura) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=chave)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill("solid", start_color=COR_H)
        cell.alignment = Alignment(horizontal="center")
        cell.border = borda
        ws.column_dimensions[get_column_letter(col)].width = largura

    # Linha 2 — legível
    for col, (_, legivel, _) in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=legivel)
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill("solid", start_color="E8F5E9")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = borda

    # Linha 3 — instrução
    instrucoes = [
        "Ex: Ceratti, Berna, Hans...",
        "Ex: Linguiça Toscana com Chimichurri Ceratti 400g",
        "Ex: Ling. Chimichurri Ceratti 400g",
        "Ex: 400",
        "g",
        "Ex: 7891234567890",
        "Ex: Linguiças Finas",
        "1",
        "direto",
        "Ex: 5  (código da Ling. Chimichurri Specialli)",
        "Opcional",
    ]
    for col, inst in enumerate(instrucoes, 1):
        cell = ws.cell(row=3, column=col, value=inst)
        cell.font = Font(italic=True, color="888888", size=8)
        cell.fill = PatternFill("solid", start_color=COR_I)
        cell.border = borda

    # Linhas de exemplo (5 produtos)
    exemplos = [
        ("Ceratti","Linguiça com Chimichurri Ceratti 400g","Ling. Chimichurri Ceratti 400g",400,"g","","Linguiças Grossas Suínas",1,"direto","5","Premium, mesmo peso e sabor"),
        ("Ceratti","Linguiça com Pimenta Biquinho Ceratti 400g","Ling. Pim. Biquinho Ceratti 400g",400,"g","","Linguiças Grossas Suínas",1,"direto","8","Premium, mesmo peso"),
        ("Berna","Salsicha Frankfurter Berna 300g","Salsicha Frankfurter Berna 300g",300,"g","","Salsichas",1,"direto","17","Artesanal, peso igual"),
        ("Seara","Linguiça Toscana Seara 600g","Ling. Toscana Seara 600g",600,"g","","Linguiças Finas",1,"indireto","1580","Industrial, peso diferente"),
        ("Seara","Linguiça Toscana Seara 5kg","Ling. Toscana Seara 5kg",5000,"g","","Linguiças Finas",0,"indireto","","Food service - não auditável"),
    ]
    for i, row in enumerate(exemplos, 4):
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = Font(size=9, color="555555", italic=True)
            cell.fill = PatternFill("solid", start_color="FFFDE7")
            cell.border = borda

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 25
    ws.freeze_panes = "A4"

    # Aba 2: Referência de categorias e códigos Specialli
    ws2 = wb.create_sheet("Referência")
    ws2["A1"] = "CATEGORIAS VÁLIDAS"
    ws2["A1"].font = Font(bold=True, color="FFFFFF")
    ws2["A1"].fill = PatternFill("solid", start_color=COR_H)
    categorias = [
        "Linguiças Finas","Linguiças Grossas Suínas","Linguiças Grossas Bovinas",
        "Linguiças Defumadas","Salsichas","Fatiados","Mortadela E Pastrami",
        "Molhos","Caixas Compostas / Kits","Massa Luinguiça"
    ]
    for i, c in enumerate(categorias, 2):
        ws2.cell(row=i, column=1, value=c).font = Font(size=9)

    ws2["C1"] = "TIPOS DE RELAÇÃO"
    ws2["C1"].font = Font(bold=True, color="FFFFFF")
    ws2["C1"].fill = PatternFill("solid", start_color=COR_H)
    ws2.cell(row=2, column=3, value="direto — mesmo produto, mesma ocasião de compra")
    ws2.cell(row=3, column=3, value="indireto — categoria próxima, peso diferente ou industrial")

    ws2["E1"] = "AUDITÁVEL"
    ws2["E1"].font = Font(bold=True, color="FFFFFF")
    ws2["E1"].fill = PatternFill("solid", start_color=COR_H)
    ws2.cell(row=2, column=5, value="1 = Auditável (entra nas análises de preço e share)")
    ws2.cell(row=3, column=5, value="0 = Não auditável (food service, industrial de massa, registrado para contexto)")

    for col in ["A","C","E"]:
        ws2.column_dimensions[col].width = 55

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf